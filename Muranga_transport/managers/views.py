from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Count, Sum, Q
from drivers.models import DriverProfile, Vehicle, MaintenanceRequest, RequestIssue, KENYA_LICENSE_CLASSES
from mechanics.models import MechanicProfile, MechanicTask
from core.models import User, RegistrationRequest
from django.views.decorators.http import require_POST
from django.contrib.auth.hashers import make_password
from datetime import timedelta
from django.forms import modelformset_factory
import random
import string
import uuid
import re
import io
import csv
import datetime
from core.utils import generate_unique_id, get_greeting
from drivers.models import SupportRequest
from mechanics.models import MechanicSupportRequest, RepairInvoice
from .models import SupportResponse, ManagerProfile, DriverAssignment, VehicleAssignment
import logging
from .forms import ManagerProfileForm, PasswordChangeForm, MechanicTaskForm, DriverAssignmentForm, VehicleAssignmentForm, InvoiceStatusForm, InvoiceApprovalForm, InvoiceRejectionForm
from django.contrib.auth import update_session_auth_hash
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl import Workbook
from django.apps import apps
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth, TruncYear
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer, Image as PDFImage
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.contrib.staticfiles import finders
from reportlab.lib.utils import Image as ReportlabImage  # Explicitly import the Image class
from django.http import JsonResponse
from django.http import HttpResponseBadRequest


logger = logging.getLogger(__name__)

def get_user(request):
    return request.user

def generate_unique_id(prefix, length=8):
    chars = string.ascii_uppercase + string.digits
    while True:
        uid = f"{prefix}_{''.join(random.choices(chars, k=length))}"
        if prefix == 'DRV' and not DriverProfile.objects.filter(driver_id=uid).exists():
            return uid
        if prefix == 'MEC' and not MechanicProfile.objects.filter(mechanic_id=uid).exists():
            return uid
        if prefix == 'TASK' and not MechanicTask.objects.filter(unique_task_id=uid).exists():
            return uid

def get_greeting():
    hour = timezone.localtime().hour
    return "Good morning" if 5 <= hour < 12 else "Good afternoon" if 12 <= hour < 17 else "Good evening"

@login_required
def dashboard(request):
    return render(request, "managers/dashboard.html", {
        "greeting": get_greeting(),
        "name": request.user.get_full_name() or request.user.username
    })

@login_required
def maintenance(request):
    status = request.GET.get('status', 'pending')
    priority = request.GET.get('priority', 'all')
    vehicle = request.GET.get('vehicle', 'all')
    query = request.GET.get('q', '')

    qs = MaintenanceRequest.objects.all().select_related('driver__user', 'vehicle').prefetch_related('issues').order_by('-submitted_at')
    if status != 'all': qs = qs.filter(status=status)
    if priority != 'all': qs = qs.filter(issues__priority=priority).distinct()
    if vehicle != 'all': qs = qs.filter(vehicle__number_plate=vehicle)
    if query:
        qs = qs.filter(
            Q(id__icontains=query) |
            Q(vehicle__number_plate__icontains=query) |
            Q(driver__user__first_name__icontains=query) |
            Q(driver__user__last_name__icontains=query) |
            Q(issues__title__icontains=query) |
            Q(issues__description__icontains=query)
        ).distinct()

    page_obj = Paginator(qs, 5).get_page(request.GET.get('page'))

    return render(request, 'managers/maintenance_requests.html', {
        "greeting": get_greeting(),
        "name": request.user.get_full_name() or request.user.username,
        "maintenance_requests": page_obj,
        "selected_status": status,
        "selected_priority": priority,
        "selected_vehicle": vehicle,
        "query": query,
        "page_obj": page_obj,
        "status_choices": MaintenanceRequest.STATUS_CHOICES,
        "priority_choices": MaintenanceRequest.PRIORITY_CHOICES,
        "vehicles": Vehicle.objects.all()
    })

@login_required
def request_detail(request, id):
    request_obj = get_object_or_404(MaintenanceRequest, pk=id)
    can_unapprove = False
    rejection_reason = request_obj.assigned_task.rejection_reason if hasattr(request_obj, 'assigned_task') and request_obj.assigned_task.status == 'rejected' else None

    if request_obj.status == 'approved' and request_obj.approved_at:
        time_diff = timezone.now() - request_obj.approved_at
        can_unapprove = time_diff <= timedelta(hours=24)

    return render(request, 'managers/request_detail.html', {
        'request_obj': request_obj,
        'can_unapprove': can_unapprove,
        'rejection_reason': rejection_reason,
        'name': request.user.get_full_name(),
        'greeting': get_greeting(),
    })

@login_required
def approve_request(request, id):
    req = get_object_or_404(MaintenanceRequest, id=id)
    priority_choices = MaintenanceRequest.PRIORITY_CHOICES

    if request.method == 'POST':
        for issue in req.issues.all():
            cost_field = f'revised_cost_{issue.id}'
            priority_field = f'priority_{issue.id}'

            revised_cost = request.POST.get(cost_field)
            new_priority = request.POST.get(priority_field)

            if revised_cost:
                try:
                    issue.cost_estimate = float(revised_cost)
                except ValueError:
                    messages.error(request, f"Invalid cost entered for issue '{issue.title}'.")
                    return redirect('managers:approve_request', id=id)

            if new_priority and new_priority in dict(priority_choices):
                issue.priority = new_priority

            issue.save()

        comment = request.POST.get('comment', '').strip()
        timestamp = timezone.now().strftime('%Y-%m-%d %H:%M')

        if comment:
            req.last_update = f"{request.user.get_full_name()} (Manager) on {timestamp}: {comment}"
        else:
            req.last_update = f"Approved by {request.user.get_full_name()} on {timestamp}"

        req.status = 'approved'
        req.approved_by = request.user
        req.approved_at = timezone.now()
        req.save()

        messages.success(request, f"Request #{req.id} has been approved.")
        return redirect('managers:maintenance')

    can_unapprove = False
    if req.status == 'approved' and req.approved_at:
        delta = timezone.now() - req.approved_at
        can_unapprove = delta <= timedelta(hours=24)

    return render(request, 'managers/approve_request.html', {
        'request_obj': req,
        'priority_choices': priority_choices,
        'can_unapprove': can_unapprove,
    })

@login_required
def unapprove_request(request, id):
    req = get_object_or_404(MaintenanceRequest, id=id)
    if req.status != 'approved':
        messages.error(request, "Only approved requests can be unapproved.")
    elif not req.approved_at or timezone.now() > req.approved_at + timedelta(hours=24):
        messages.error(request, "You can only unapprove within 24 hours.")
    else:
        req.status = 'pending'
        req.approved_by = None
        req.approved_at = None
        req.last_update = f"Unapproved by {request.user.get_full_name()} on {timezone.now().strftime('%Y-%m-%d %H:%M')}"
        req.save()
        messages.success(request, f"Request #{id} unapproved successfully.")
    return redirect('managers:maintenance')

@login_required
def deny_request(request, id):
    req = get_object_or_404(MaintenanceRequest, id=id)
    if request.method == 'POST':
        reason = request.POST.get('deny_reason', '').strip()
        if reason:
            req.status = 'denied'
            req.last_update = f"Denied by {request.user.get_full_name()} on {timezone.now().strftime('%Y-%m-%d %H:%M')}: {reason}"
            req.save()
            messages.success(request, f"Request #{id} denied.")
        else:
            messages.error(request, "Denial reason is required.")
            return redirect('managers:request_detail', id=id)
    return redirect('managers:maintenance')

@login_required
def assign_mechanic(request, id):
    req = get_object_or_404(MaintenanceRequest, id=id)
    if req.status not in ['approved', 'rejected']:
        messages.error(request, "Request must be approved or rejected to assign a mechanic.")
        return redirect('managers:maintenance')
    if hasattr(req, 'assigned_task') and req.assigned_task.status not in ['rejected', 'completed']:
        messages.error(request, "Task is already assigned and not rejected or completed.")
        return redirect('managers:maintenance')

    query = request.GET.get('q', '')
    mechanics = MechanicProfile.objects.filter(status='active').order_by('full_name')
    if query:
        mechanics = mechanics.filter(Q(full_name__icontains=query) | Q(mechanic_id__icontains=query))

    page_obj = Paginator(mechanics, 5).get_page(request.GET.get('page'))

    if request.method == 'POST':
        mech = get_object_or_404(MechanicProfile, id=request.POST.get('mechanic_id'))
        if hasattr(req, 'assigned_task') and req.assigned_task.status == 'rejected':
            task = req.assigned_task
            task.mechanic = mech
            task.status = 'pending'
            task.rejection_reason = None
            task.notes = f"Reassigned to {mech.full_name} on {timezone.now().strftime('%Y-%m-%d %H:%M')}"
            task.save()
        else:
            MechanicTask.objects.create(
                maintenance_request=req,  # Changed from maintenance_request_id=req
                mechanic=mech,
                status='pending',
                priority=req.issues.first().priority if req.issues.exists() else 'medium',
                notes=f"Assigned to {mech.full_name} on {timezone.now().strftime('%Y-%m-%d %H:%M')}"
            )
        req.mechanic = mech
        req.status = 'pending'
        req.last_update = f"Assigned to {mech.full_name} on {timezone.now().strftime('%Y-%m-%d %H:%M')}"
        req.save()
        messages.success(request, f"Assigned mechanic to request #{id}")
        return redirect('managers:maintenance')

    return render(request, 'managers/assign_mechanic.html', {
        "greeting": get_greeting(),
        "name": request.user.get_full_name() or request.user.username,
        "request_obj": req,
        "mechanics": page_obj,
        "query": query,
        "page_obj": page_obj
    })

@login_required
def applications_view(request):
    query = request.GET.get('q', '')
    role = request.GET.get('role', 'all')
    status = request.GET.get('status', 'all')
    license_class = request.GET.get('license', 'all')
    specialty = request.GET.get('specialty', 'all')

    reg_qs = RegistrationRequest.objects.filter(is_approved=False).order_by('-created_at')
    if role != 'all': reg_qs = reg_qs.filter(role=role)
    if query: reg_qs = reg_qs.filter(Q(full_name__icontains=query) | Q(email__icontains=query) | Q(phone_number__icontains=query))

    drivers = DriverProfile.objects.select_related('user').all().order_by('user__username')
    if status != 'all': drivers = drivers.filter(status=status)
    if license_class != 'all': drivers = drivers.filter(license_class=license_class)

    mechanics = MechanicProfile.objects.select_related('user').all().order_by('full_name')
    if status != 'all': mechanics = mechanics.filter(status=status)
    if specialty != 'all': mechanics = mechanics.filter(specialization=specialty)

    page_obj = Paginator(reg_qs, 5).get_page(request.GET.get('page'))

    return render(request, 'managers/applications.html', {
        "greeting": get_greeting(),
        "name": request.user.get_full_name() or request.user.username,
        "registration_requests": page_obj,
        "drivers": drivers,
        "mechanics": mechanics,
        "query": query,
        "role_filter": role,
        "status_filter": status,
        "license_filter": license_class,
        "specialty_filter": specialty
    })

@login_required
@require_POST
def approve_registration(request, id):
    reg = get_object_or_404(RegistrationRequest, id=id)

    if reg.is_approved:
        messages.error(request, 'Registration already approved.')
        return redirect('managers:applications')

    if User.objects.filter(email=reg.email).exists():
        messages.error(request, 'User with this email already exists.')
        return redirect('managers:applications')

    try:
        full_name_parts = reg.full_name.strip().split()
        first_name = full_name_parts[0]
        last_name = " ".join(full_name_parts[1:]) if len(full_name_parts) > 1 else ''
        username = f"{reg.role}_{uuid.uuid4().hex[:8]}"
        user = User.objects.create(
            username=username,
            email=reg.email,
            phone_number=reg.phone_number,
            first_name=first_name,
            last_name=last_name,
            role=reg.role,
            dl_no=reg.drivers_license if reg.role == 'driver' else None,
            id_no=reg.id_number if reg.id_number else None,
            location=reg.location if reg.location else None,
            is_active=True,
            password=reg.password
        )

        if reg.role == 'driver':
            DriverProfile.objects.create(
                user=user,
                driver_id=generate_unique_id('DRV'),
                phone_number=reg.phone_number,
                license_class=reg.license_class,
                experience_years=reg.experience_years,
                department=reg.department if reg.department else None,
                supervisor=reg.supervisor if reg.supervisor else None
            )
        elif reg.role == 'mechanic':
            MechanicProfile.objects.create(
                user=user,
                mechanic_id=generate_unique_id('MEC'),
                full_name=reg.full_name,
                phone=reg.phone_number,
                email=reg.email,
                experience_years=reg.experience_years,
                location=reg.location if reg.location else None,
                specialization=reg.specialization if reg.specialization else None
            )

        reg.delete()
        messages.success(request, f"{reg.full_name} approved and moved to active users.")

    except Exception as e:
        messages.error(request, f"Approval failed: {str(e)}")
        return redirect('managers:applications')

    return redirect('managers:applications')

@login_required
@require_POST
def deny_registration(request, id):
    if request.method == 'POST':
        reg = get_object_or_404(RegistrationRequest, id=id)
        reason = request.POST.get('deny_reason', '').strip()
        print(f"[INFO] Denied {reg.full_name}. Reason: {reason if reason else 'No reason provided'}")
        reg.delete()
        messages.success(request, f"{reg.full_name}'s registration request denied and deleted.")
    return redirect('managers:applications')

@login_required
def delete_driver(request, id):
    driver = get_object_or_404(DriverProfile, id=id)
    user = driver.user
    user.delete()
    messages.success(request, f"Driver {driver.driver_id} deleted.")
    return redirect('managers:applications')

@login_required
def delete_mechanic(request, id):
    mechanic = get_object_or_404(MechanicProfile, id=id)
    user = mechanic.user
    user.delete()
    messages.success(request, f"Mechanic {mechanic.mechanic_id} deleted.")
    return redirect('managers:applications')

@login_required
def personnel_management(request):
    driver_query = request.GET.get('driver_search', '')
    mechanic_query = request.GET.get('mechanic_search', '')

    drivers_qs = DriverProfile.objects.select_related('user').order_by('user__username')
    if driver_query:
        drivers_qs = drivers_qs.filter(
            Q(user__first_name__icontains=driver_query) |
            Q(user__last_name__icontains=driver_query) |
            Q(user__email__icontains=driver_query)
        )

    mechanics_qs = MechanicProfile.objects.all().order_by('full_name')
    if mechanic_query:
        mechanics_qs = mechanics_qs.filter(
            Q(full_name__icontains=mechanic_query) |
            Q(email__icontains=mechanic_query) |
            Q(phone__icontains=mechanic_query)
        )

    drivers_paged = Paginator(drivers_qs, 5).get_page(request.GET.get('driver_page'))
    mechanics_paged = Paginator(mechanics_qs, 5).get_page(request.GET.get('mechanic_page'))

    return render(request, 'managers/personnel_management.html', {
        'greeting': get_greeting(),
        'name': request.user.get_full_name() or request.user.username,
        'drivers': drivers_paged,
        'mechanics': mechanics_paged,
        'driver_query': driver_query,
        'mechanic_query': mechanic_query,
    })

@login_required
def add_user_form(request):
    if request.method == 'POST':
        role = request.POST.get('role')
        first_name = request.POST.get('firstName')
        last_name = request.POST.get('lastName')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone')
        id_number = request.POST.get('idNumber')
        location = request.POST.get('location')
        experience_years = request.POST.get('experienceYears')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirmPassword')
        drivers_license = request.POST.get('driversLicense')
        license_classes = request.POST.getlist('licenseClass')
        department = request.POST.get('department')
        supervisor = request.POST.get('supervisor')
        specialization = request.POST.get('specialization')

        if not all([role, first_name, last_name, email, phone_number, experience_years, password, confirm_password]):
            messages.error(request, 'Required fields missing.')
            return render(request, 'managers/add_user.html', {'KENYA_LICENSE_CLASSES': KENYA_LICENSE_CLASSES})

        if role not in ['driver', 'mechanic']:
            messages.error(request, 'Invalid role selected.')
            return render(request, 'managers/add_user.html', {'KENYA_LICENSE_CLASSES': KENYA_LICENSE_CLASSES})

        if password != confirm_password:
            messages.error(request, 'Passwords do not match.')
            return render(request, 'managers/add_user.html', {'KENYA_LICENSE_CLASSES': KENYA_LICENSE_CLASSES})

        if len(password) < 8 or not re.search(r'[A-Z]', password) or not re.search(r'[a-z]', password) or \
           not re.search(r'\d', password) or not re.search(r'[@$!%*?&]', password):
            messages.error(request, 'Password must be strong: min 8 chars, upper, lower, number & special.')
            return render(request, 'managers/add_user.html', {'KENYA_LICENSE_CLASSES': KENYA_LICENSE_CLASSES})

        if User.objects.filter(email=email).exists():
            messages.error(request, 'Email already used.')
            return render(request, 'managers/add_user.html', {'KENYA_LICENSE_CLASSES': KENYA_LICENSE_CLASSES})

        if User.objects.filter(phone_number=phone_number).exists():
            messages.error(request, 'Phone number already used.')
            return render(request, 'managers/add_user.html', {'KENYA_LICENSE_CLASSES': KENYA_LICENSE_CLASSES})

        if User.objects.filter(id_no=id_number).exists():
            messages.error(request, 'ID number already used.')
            return render(request, 'managers/add_user.html', {'KENYA_LICENSE_CLASSES': KENYA_LICENSE_CLASSES})

        if role == 'driver':
            if not drivers_license:
                messages.error(request, 'Driver’s license number is required.')
                return render(request, 'managers/add_user.html', {'KENYA_LICENSE_CLASSES': KENYA_LICENSE_CLASSES})
            if not license_classes:
                messages.error(request, 'At least one license class is required.')
                return render(request, 'managers/add_user.html', {'KENYA_LICENSE_CLASSES': KENYA_LICENSE_CLASSES})
            valid_codes = [code for code, _ in KENYA_LICENSE_CLASSES]
            for code in license_classes:
                if code not in valid_codes:
                    messages.error(request, f"Invalid license class: {code}.")
                    return render(request, 'managers/add_user.html', {'KENYA_LICENSE_CLASSES': KENYA_LICENSE_CLASSES})

        try:
            full_name = f"{first_name} {last_name}".strip()
            username = f"{role}_{uuid.uuid4().hex[:8]}"
            user = User.objects.create(
                username=username,
                email=email,
                phone_number=phone_number,
                first_name=first_name,
                last_name=last_name,
                role=role,
                dl_no=drivers_license if role == 'driver' else '',
                id_no=id_number,
                location=location,
                is_active=True
            )
            user.set_password(password)
            user.save()

            if role == 'driver':
                DriverProfile.objects.create(
                    user=user,
                    driver_id=generate_unique_id('DRV'),
                    phone_number=phone_number,
                    license_class=','.join(license_classes),
                    experience_years=int(experience_years),
                    department=department,
                    supervisor=supervisor
                )
            else:
                MechanicProfile.objects.create(
                    user=user,
                    mechanic_id=generate_unique_id('MEC'),
                    full_name=full_name,
                    phone=phone_number,
                    email=email,
                    experience_years=int(experience_years),
                    location=location,
                    specialization=specialization,
                    status='active'
                )

            messages.success(request, f"{role.capitalize()} '{full_name}' added successfully.")
            return redirect('managers:personnel')

        except Exception as e:
            print(f"[ERROR] User creation failed: {e}")
            messages.error(request, 'Something went wrong while adding the user.')
            return render(request, 'managers/add_user.html', {'KENYA_LICENSE_CLASSES': KENYA_LICENSE_CLASSES})

    return render(request, 'managers/add_user.html', {'KENYA_LICENSE_CLASSES': KENYA_LICENSE_CLASSES})

@login_required
def job_management(request):
    greeting = get_greeting()

    if request.method == "POST":
        if "task_submit" in request.POST:
            task_form = MechanicTaskForm(request.POST)
            driver_assignment_form = DriverAssignmentForm()
            vehicle_assignment_form = VehicleAssignmentForm()
            if task_form.is_valid():
                task = task_form.save(commit=False)
                task.assigned_at = timezone.now()
                task.save()
                messages.success(request, f"Task {task.unique_task_id} created successfully.")
                return redirect("managers:job_management")
            else:
                messages.error(request, "Please fix the errors in the task form.")
        elif "driver_assignment_submit" in request.POST:
            driver_assignment_form = DriverAssignmentForm(request.POST)
            task_form = MechanicTaskForm()
            vehicle_assignment_form = VehicleAssignmentForm()
            if driver_assignment_form.is_valid():
                driver_assignment_form.save()
                messages.success(request, "Driver assignment created successfully.")
                return redirect("managers:job_management")
            else:
                messages.error(request, "Please fix the errors in the driver assignment form.")
        elif "vehicle_assignment_submit" in request.POST:
            vehicle_assignment_form = VehicleAssignmentForm(request.POST)
            task_form = MechanicTaskForm()
            driver_assignment_form = DriverAssignmentForm()
            if vehicle_assignment_form.is_valid():
                vehicle_assignment_form.save()
                messages.success(request, "Vehicle assigned successfully.")
                return redirect("managers:job_management")
            else:
                messages.error(request, "Please fix the errors in the vehicle assignment form.")
    else:
        task_form = MechanicTaskForm()
        driver_assignment_form = DriverAssignmentForm()
        vehicle_assignment_form = VehicleAssignmentForm()

    return render(request, "managers/job_management.html", {
        "task_form": task_form,
        "driver_assignment_form": driver_assignment_form,
        "vehicle_assignment_form": vehicle_assignment_form,
        "greeting": greeting,
    })

@login_required
def mechanic_tasks(request):
    user = get_user(request)
    tasks = MechanicTask.objects.filter(manager=user).select_related('mechanic__user').order_by('-created_at')
    search_query = request.GET.get('q', '')
    status_filter = request.GET.get('status', 'all')

    if search_query:
        tasks = tasks.filter(
            Q(issue_title__icontains=search_query) |
            Q(issue_description__icontains=search_query) |
            Q(vehicle_number_plate__icontains=search_query) |
            Q(mechanic__user__first_name__icontains=search_query) |
            Q(mechanic__user__last_name__icontains=search_query)
        )
    
    if status_filter != 'all':
        tasks = tasks.filter(status=status_filter)

    paginator = Paginator(tasks, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'managers/mechanic_tasks.html', {
        'greeting': get_greeting(),
        'user': user,
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
    })

@login_required
def driver_assignments(request):
    user = get_user(request)
    assignments = DriverAssignment.objects.filter(manager=user).select_related('driver').order_by('-created_at')
    search_query = request.GET.get('q', '')
    status_filter = request.GET.get('status', 'all')

    if search_query:
        assignments = assignments.filter(
            Q(title__icontains=search_query) |
            Q(destination__icontains=search_query) |
            Q(vehicle_number_plate__icontains=search_query) |
            Q(driver__first_name__icontains=search_query) |
            Q(driver__last_name__icontains=search_query)
        )
    
    if status_filter != 'all':
        assignments = assignments.filter(status=status_filter)

    paginator = Paginator(assignments, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'managers/driver_assignments.html', {
        'greeting': get_greeting(),
        'user': user,
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
    })

@login_required
def repair_invoices(request):
    RepairInvoice = apps.get_model('mechanics', 'RepairInvoice')
    invoices = RepairInvoice.objects.select_related('mechanic_task__mechanic').all().order_by('-date_of_service')

    status = request.GET.get('status')
    mechanic_id = request.GET.get('mechanic')
    search_query = request.GET.get('search')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if status and status != 'all':
        invoices = invoices.filter(status__iexact=status)

    if mechanic_id and mechanic_id != 'all':
        invoices = invoices.filter(mechanic_task__mechanic__id=mechanic_id)

    if search_query:
        invoices = invoices.filter(
            Q(task_unique_id__icontains=search_query) |
            Q(issues__icontains=search_query) |
            Q(vehicle_number__icontains=search_query)
        )

    if start_date:
        invoices = invoices.filter(date_of_service__gte=start_date)
    if end_date:
        invoices = invoices.filter(date_of_service__lte=end_date)

    paginator = Paginator(invoices, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    total_paid = RepairInvoice.objects.filter(status='paid').aggregate(Sum('total_cost'))['total_cost__sum'] or 0
    pending_approval = RepairInvoice.objects.filter(status='pending').aggregate(Sum('total_cost'))['total_cost__sum'] or 0
    rejected_invoices = RepairInvoice.objects.filter(status='rejected').aggregate(Sum('total_cost'))['total_cost__sum'] or 0

    pending_count = RepairInvoice.objects.filter(status='pending').count()
    rejected_count = RepairInvoice.objects.filter(status='rejected').count()
    mechanics = MechanicProfile.objects.filter(status='active').order_by('full_name')

    context = {
        'invoices': page_obj,
        'page_obj': page_obj,
        'mechanics': mechanics,
        'total_paid': total_paid,
        'pending_approval': pending_approval,
        'rejected_invoices': rejected_invoices,
        'pending_count': pending_count,
        'rejected_count': rejected_count,
        'status_choices': RepairInvoice.STATUS_CHOICES,
    }
    return render(request, 'managers/repair_invoices.html', context)

@login_required
def approve_invoice(request, invoice_id):
    RepairInvoice = apps.get_model('mechanics', 'RepairInvoice')
    invoice = get_object_or_404(RepairInvoice, pk=invoice_id)
    if request.method == 'POST':
        form = InvoiceApprovalForm(request.POST)
        if form.is_valid():
            invoice.status = 'approved'
            invoice.action_timestamp = timezone.now()
            invoice.rejected_by = None
            invoice.rejection_reason = ''
            invoice.rejection_comment = ''
            invoice.save()
            messages.success(request, f"Invoice #{invoice.task_unique_id} approved.")
            return redirect('managers:repair_invoices')
    else:
        form = InvoiceApprovalForm()
    return render(request, 'managers/approve_invoice.html', {'invoice': invoice, 'form': form})

@login_required
def reject_invoice(request, invoice_id):
    RepairInvoice = apps.get_model('mechanics', 'RepairInvoice')
    invoice = get_object_or_404(RepairInvoice, pk=invoice_id)
    if request.method == 'POST':
        form = InvoiceRejectionForm(request.POST)
        if form.is_valid():
            invoice.status = 'rejected'
            invoice.rejected_by = request.user
            invoice.rejection_reason = form.cleaned_data['reason']
            invoice.rejection_comment = form.cleaned_data['notes']
            invoice.action_timestamp = timezone.now()
            invoice.save()
            messages.warning(request, f"Invoice #{invoice.task_unique_id} rejected.")
            return redirect('managers:repair_invoices')
    else:
        form = InvoiceRejectionForm()
    return render(request, 'managers/reject_invoice.html', {'invoice': invoice, 'form': form})

@login_required
def unapprove_invoice(request, invoice_id):
    RepairInvoice = apps.get_model('mechanics', 'RepairInvoice')
    invoice = get_object_or_404(RepairInvoice, pk=invoice_id)
    if invoice.status != 'approved':
        messages.error(request, "Invoice is not approved.")
    elif timezone.now() - invoice.action_timestamp > timedelta(hours=48):
        messages.error(request, "Unapproval window expired (48 hours).")
    else:
        invoice.status = 'pending'
        invoice.action_timestamp = timezone.now()
        invoice.save()
        messages.success(request, f"Invoice #{invoice.task_unique_id} moved back to Pending.")
    return redirect('managers:repair_invoices')

@login_required
def unreject_invoice(request, invoice_id):
    RepairInvoice = apps.get_model('mechanics', 'RepairInvoice')
    invoice = get_object_or_404(RepairInvoice, pk=invoice_id)
    if invoice.status != 'rejected':
        messages.error(request, "Invoice is not rejected.")
    elif timezone.now() - invoice.action_timestamp > timedelta(hours=48):
        messages.error(request, "Unrejection window expired (48 hours).")
    else:
        invoice.status = 'pending'
        invoice.rejected_by = None
        invoice.rejection_reason = ''
        invoice.rejection_comment = ''
        invoice.action_timestamp = timezone.now()
        invoice.save()
        messages.success(request, f"Invoice #{invoice.task_unique_id} moved back to Pending.")
    return redirect('managers:repair_invoices')

def export_excel(invoices):
    wb = Workbook()
    ws = wb.active
    ws.title = "Repair Invoices"

    headers = ['Invoice #', 'Date of Service', 'Vehicle', 'Mechanic', 'Issues', 'Total Cost', 'Status']
    ws.append(headers)

    for inv in invoices:
        ws.append([
            inv.task_unique_id,
            inv.date_of_service.strftime('%Y-%m-%d'),
            inv.vehicle_number,
            inv.mechanic_task.mechanic_profile.full_name if inv.mechanic_task and inv.mechanic_task.mechanic_profile else '',
            inv.issues,
            float(inv.total_cost),
            inv.get_status_display(),
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"repair_invoices_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def export_pdf(invoices):
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 40

    p.setFont("Helvetica-Bold", 14)
    p.drawString(200, y, "Repair Invoices Report")
    y -= 30

    p.setFont("Helvetica", 10)
    headers = ['Invoice #', 'Date', 'Vehicle', 'Mechanic', 'Issues', 'Amount', 'Status']
    col_widths = [60, 60, 60, 80, 100, 50, 50]

    for i, header in enumerate(headers):
        p.drawString(30 + sum(col_widths[:i]), y, header)
    y -= 20

    for inv in invoices:
        row = [
            inv.task_unique_id,
            inv.date_of_service.strftime('%Y-%m-%d'),
            inv.vehicle_number,
            inv.mechanic_task.mechanic_profile.full_name if inv.mechanic_task and inv.mechanic_task.mechanic_profile else '',
            (inv.issues or '')[:30],
            f"{inv.total_cost:.2f}",
            inv.get_status_display(),
        ]
        for i, item in enumerate(row):
            p.drawString(30 + sum(col_widths[:i]), y, str(item))
        y -= 20
        if y < 50:
            p.showPage()
            p.setFont("Helvetica", 10)
            y = height - 40

    p.save()
    buffer.seek(0)

    filename = f"repair_invoices_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def export_invoices(request, format):
    RepairInvoice = apps.get_model('mechanics', 'RepairInvoice')
    invoices = RepairInvoice.objects.select_related('mechanic_task__mechanic_profile').all().order_by('-date_of_service')

    status = request.GET.get('status')
    mechanic_id = request.GET.get('mechanic')
    search_query = request.GET.get('search')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if status and status != 'all':
        invoices = invoices.filter(status__iexact=status)
    if mechanic_id and mechanic_id != 'all':
        invoices = invoices.filter(mechanic_task__mechanic_profile__id=mechanic_id)
    if search_query:
        invoices = invoices.filter(
            Q(task_unique_id__icontains=search_query) |
            Q(issues__icontains=search_query) |
            Q(vehicle_number__icontains=search_query)
        )
    if start_date:
        invoices = invoices.filter(date_of_service__gte=start_date)
    if end_date:
        invoices = invoices.filter(date_of_service__lte=end_date)

    if format == 'excel':
        return export_excel(invoices)
    elif format == 'pdf':
        return export_pdf(invoices)
    else:
        return HttpResponse("Invalid export format", status=400)

def export_invoices_pdf(request):
    RepairInvoice = apps.get_model('mechanics', 'RepairInvoice')
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="repair_invoices.pdf"'

    p = canvas.Canvas(response)
    invoices = RepairInvoice.objects.all().order_by('-date_of_service')

    y = 800
    p.drawString(100, y, "Repair Invoices Report")
    y -= 30
    for invoice in invoices:
        line = f"#{invoice.task_unique_id} - {invoice.mechanic_task.mechanic_profile.full_name if invoice.mechanic_task and invoice.mechanic_task.mechanic_profile else ''} - Ksh {invoice.total_cost} - {invoice.status}"
        p.drawString(100, y, line)
        y -= 20
        if y < 50:
            p.showPage()
            y = 800

    p.showPage()
    p.save()
    return response

@login_required
def reports(request):
    # Maintenance Stats 
    total_requests = MaintenanceRequest.objects.count()
    pending_requests = MaintenanceRequest.objects.filter(status='pending').count()
    approved_requests = MaintenanceRequest.objects.filter(status='approved').count()
    assigned_requests = MaintenanceRequest.objects.filter(mechanic__isnull=False).count()
    completed_requests = MaintenanceRequest.objects.filter(status='completed').count()

    # High-priority requests based on related issues
    high_priority_requests = MaintenanceRequest.objects.filter(
        issues__priority='high'
    ).distinct().count()

    # Recent maintenance requests (displayed in Maintenance tab)
    requests_qs = MaintenanceRequest.objects.select_related(
        'driver__user', 'vehicle'
    ).order_by('-submitted_at')[:10]

    #  Mechanics Stats
    mechanic_stats = (
        MechanicProfile.objects.annotate(
            total_tasks=Count('assigned_requests', distinct=True),
            accepted_tasks=Count(
                'assigned_requests', filter=Q(assigned_requests__status='approved'), distinct=True
            ),
            completed_tasks=Count(
                'assigned_requests', filter=Q(assigned_requests__status='completed'), distinct=True
            ),
            invoices_sent=Count('mechanictask__invoice', distinct=True),
            pending_invoices=Count(
                'mechanictask__invoice',
                filter=Q(mechanictask__invoice__status='pending'),
                distinct=True,
            ),
            approved_invoices=Count(
                'mechanictask__invoice',
                filter=Q(mechanictask__invoice__status='approved'),
                distinct=True,
            ),
            paid_invoices=Count(
                'mechanictask__invoice',
                filter=Q(mechanictask__invoice__status='paid'),
                distinct=True,
            ),
        )
        .values(
            'mechanic_id',
            'full_name',
            'total_tasks',
            'accepted_tasks',
            'completed_tasks',
            'invoices_sent',
            'pending_invoices',
            'approved_invoices',
            'paid_invoices',
        )
        .order_by('-total_tasks')
    )

    # Driver Stats
    driver_stats = (
        DriverProfile.objects.annotate(
            total_requests=Count('maintenancerequest', distinct=True),
            pending_requests=Count(
                'maintenancerequest',
                filter=Q(maintenancerequest__status='pending'),
                distinct=True,
            ),
            approved_requests=Count(
                'maintenancerequest',
                filter=Q(maintenancerequest__status='approved'),
                distinct=True,
            ),
            completed_requests=Count(
                'maintenancerequest',
                filter=Q(maintenancerequest__status='completed'),
                distinct=True,
            ),
        )
        .values(
            'driver_id',
            'user__first_name',
            'user__last_name',
            'total_requests',
            'pending_requests',
            'approved_requests',
            'completed_requests',
        )
        .order_by('-total_requests')
    )

    #  Vehicle Stats
    vehicle_stats = (
        Vehicle.objects.annotate(
            total_repairs=Count('maintenancerequest', distinct=True)
        )
        .values('number_plate', 'make', 'model', 'total_repairs')
        .order_by('-total_repairs')
    )

    # Invoices Stats 
    invoices_qs = RepairInvoice.objects.select_related('mechanic_task__mechanic').order_by('-created_at')[:10]

    context = {
        # Main stats
        'total_requests': total_requests,
        'pending_requests': pending_requests,
        'approved_requests': approved_requests,
        'assigned_requests': assigned_requests,
        'completed_requests': completed_requests,
        'high_priority_requests': high_priority_requests,

        # Tables
        'requests': requests_qs,
        'mechanic_stats': mechanic_stats,
        'driver_stats': driver_stats,
        'vehicle_stats': vehicle_stats,
        'invoices': invoices_qs,

        'time_range': request.GET.get('time_range', 'monthly'),
    }
    return render(request, 'managers/reports.html', context)


@login_required
def reports_chart_data(request):
    """AJAX endpoint for real-time chart updates."""
    time_range = request.GET.get('time_range', 'day')

    # Maintenance Chart (pie)
    maintenance_status = MaintenanceRequest.objects.values('status').annotate(count=Count('id'))
    maintenance_labels = [m['status'].capitalize() for m in maintenance_status]
    maintenance_data = [m['count'] for m in maintenance_status]

    # Invoice Chart (pie)
    invoice_status = RepairInvoice.objects.values('status').annotate(count=Count('id'))
    invoice_labels = [i['status'].capitalize() for i in invoice_status]
    invoice_data = [i['count'] for i in invoice_status]

    # Vehicle Repairs (bar)
    vehicle_repairs = Vehicle.objects.annotate(total_repairs=Count('maintenancerequest')).values('number_plate', 'total_repairs')
    vehicle_labels = [v['number_plate'] for v in vehicle_repairs]
    vehicle_data = [v['total_repairs'] for v in vehicle_repairs]

    return JsonResponse({
        'maintenance': {'labels': maintenance_labels, 'data': maintenance_data},
        'invoices': {'labels': invoice_labels, 'data': invoice_data},
        'vehicles': {'labels': vehicle_labels, 'data': vehicle_data},
    })

@login_required
def export_reports(request):
    report_type = request.GET.get('report_type', 'maintenance')  # maintenance, invoices, mechanics, drivers, vehicles
    export_type = request.GET.get('type', 'csv')  # csv, excel, pdf
    date_range = request.GET.get('date_range', 'all')
    time_range = request.GET.get('time_range', 'day')  # day, week, month, year

    # Base querysets
    maintenance_qs = MaintenanceRequest.objects.select_related('driver__user', 'vehicle', 'mechanic').prefetch_related('issues')
    invoices_qs = RepairInvoice.objects.select_related('mechanic_task__mechanic')
    mechanic_qs = MechanicProfile.objects.all()
    driver_qs = DriverProfile.objects.all()
    vehicle_qs = Vehicle.objects.all()

    # Apply date range filter
    if date_range != 'all':
        today = timezone.now().date()
        if date_range == 'today':
            maintenance_qs = maintenance_qs.filter(submitted_at__date=today)
            invoices_qs = invoices_qs.filter(created_at__date=today)
        elif date_range == 'week':
            maintenance_qs = maintenance_qs.filter(submitted_at__gte=timezone.now() - timedelta(days=7))
            invoices_qs = invoices_qs.filter(created_at__gte=timezone.now() - timedelta(days=7))
        elif date_range == 'month':
            maintenance_qs = maintenance_qs.filter(submitted_at__gte=timezone.now() - timedelta(days=30))
            invoices_qs = invoices_qs.filter(created_at__gte=timezone.now() - timedelta(days=30))
        elif date_range == 'year':
            maintenance_qs = maintenance_qs.filter(submitted_at__year=timezone.now().year)
            invoices_qs = invoices_qs.filter(created_at__year=timezone.now().year)

    # Export handlers
    if report_type == 'maintenance':
        return _export_maintenance(export_type, maintenance_qs)
    elif report_type == 'invoices':
        return _export_invoices(export_type, invoices_qs)
    elif report_type == 'mechanics':
        mechanic_stats = mechanic_qs.annotate(
            total_tasks=Count('mechanictask'),
            accepted_tasks=Count('mechanictask', filter=Q(mechanictask__status='in_progress') | Q(mechanictask__status='completed')),
            completed_tasks=Count('mechanictask', filter=Q(mechanictask__status='completed')),
            invoices_sent=Count('mechanictask__invoice', distinct=True),
            pending_invoices=Count('mechanictask__invoice', filter=Q(mechanictask__invoice__status='pending')),
            approved_invoices=Count('mechanictask__invoice', filter=Q(mechanictask__invoice__status='approved')),
            paid_invoices=Count('mechanictask__invoice', filter=Q(mechanictask__invoice__status='paid'))
        )
        return _export_mechanics(export_type, mechanic_stats)
    elif report_type == 'drivers':
        driver_stats = driver_qs.annotate(
            total_requests=Count('maintenancerequest'),
            pending_requests=Count('maintenancerequest', filter=Q(maintenancerequest__status='pending')),
            approved_requests=Count('maintenancerequest', filter=Q(maintenancerequest__status='approved')),
            completed_requests=Count('maintenancerequest', filter=Q(maintenancerequest__status='completed'))
        )
        return _export_drivers(export_type, driver_stats)
    elif report_type == 'vehicles':
        vehicle_stats = vehicle_qs.annotate(total_repairs=Count('maintenancerequest'))
        return _export_vehicles(export_type, vehicle_stats)
    else:
        return HttpResponse("Invalid report type", status=400)

# Export Functions
def _export_maintenance(export_type, requests):
    if export_type == 'csv':
        return _export_maintenance_csv(requests)
    elif export_type == 'excel':
        return _export_maintenance_excel(requests)
    elif export_type == 'pdf':
        return _export_maintenance_pdf(requests)

def _export_invoices(export_type, invoices):
    if export_type == 'csv':
        return _export_invoices_csv(invoices)
    elif export_type == 'excel':
        return _export_invoices_excel(invoices)
    elif export_type == 'pdf':
        return _export_invoices_pdf(invoices)

def _export_mechanics(export_type, mechanics):
    if export_type == 'csv':
        return _export_mechanics_csv(mechanics)
    elif export_type == 'excel':
        return _export_mechanics_excel(mechanics)
    elif export_type == 'pdf':
        return _export_mechanics_pdf(mechanics)

def _export_drivers(export_type, drivers):
    if export_type == 'csv':
        return _export_drivers_csv(drivers)
    elif export_type == 'excel':
        return _export_drivers_excel(drivers)
    elif export_type == 'pdf':
        return _export_drivers_pdf(drivers)

def _export_vehicles(export_type, vehicles):
    if export_type == 'csv':
        return _export_vehicles_csv(vehicles)
    elif export_type == 'excel':
        return _export_vehicles_excel(vehicles)
    elif export_type == 'pdf':
        return _export_vehicles_pdf(vehicles)

# CSV Exports
def _export_maintenance_csv(requests):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="maintenance_reports_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Request ID', 'Driver', 'Vehicle', 'Status', 'Submitted Date'])
    for req in requests:
        writer.writerow([
            req.pk,
            req.driver.user.get_full_name() or req.driver.user.username,
            req.vehicle.number_plate,
            req.get_status_display(),
            req.submitted_at.strftime('%Y-%m-%d')
        ])
    return response

def _export_invoices_csv(invoices):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="invoice_reports_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Invoice ID', 'Mechanic', 'Task ID', 'Total Cost', 'Status', 'Created Date'])
    for inv in invoices:
        writer.writerow([
            inv.id,
            inv.mechanic_task.mechanic.full_name if inv.mechanic_task.mechanic else "N/A",
            inv.mechanic_task.unique_task_id,
            inv.total_cost,
            inv.get_status_display(),
            inv.created_at.strftime('%Y-%m-%d')
        ])
    return response

def _export_mechanics_csv(mechanics):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="mechanic_reports_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Mechanic ID', 'Full Name', 'Total Tasks', 'Accepted Tasks', 'Completed Tasks', 'Invoices Sent',
                     'Pending Invoices', 'Approved Invoices', 'Paid Invoices'])
    for mech in mechanics:
        writer.writerow([
            mech['mechanic_id'],
            mech['full_name'],
            mech['total_tasks'],
            mech['accepted_tasks'],
            mech['completed_tasks'],
            mech['invoices_sent'],
            mech['pending_invoices'],
            mech['approved_invoices'],
            mech['paid_invoices']
        ])
    return response

def _export_drivers_csv(drivers):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="driver_reports_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Driver ID', 'Full Name', 'Total Requests', 'Pending Requests', 'Approved Requests', 'Completed Requests'])
    for driver in drivers:
        full_name = f"{driver['user__first_name']} {driver['user__last_name']}".strip() or "N/A"
        writer.writerow([
            driver['driver_id'],
            full_name,
            driver['total_requests'],
            driver['pending_requests'],
            driver['approved_requests'],
            driver['completed_requests']
        ])
    return response

def _export_vehicles_csv(vehicles):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="vehicle_reports_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Number Plate', 'Make', 'Model', 'Total Repairs'])
    for vehicle in vehicles:
        writer.writerow([
            vehicle['number_plate'],
            vehicle['make'],
            vehicle['model'],
            vehicle['total_repairs']
        ])
    return response

# Excel Exports (with styling and logo)
def _export_maintenance_excel(requests):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maintenance Reports"
    _add_logo_and_title(ws, "Maintenance Report")
    _add_table(ws, ['Request ID', 'Driver', 'Vehicle', 'Status', 'Submitted Date'], requests,
               lambda req: [req.pk, req.driver.user.get_full_name() or req.driver.user.username,
                            req.vehicle.number_plate, req.get_status_display(), req.submitted_at.strftime('%Y-%m-%d')])
    return _finalize_excel(wb, "maintenance_reports")

def _export_invoices_excel(invoices):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Reports"
    _add_logo_and_title(ws, "Invoice Report")
    _add_table(ws, ['Invoice ID', 'Mechanic', 'Task ID', 'Total Cost', 'Status', 'Created Date'], invoices,
               lambda inv: [inv.id, inv.mechanic_task.mechanic.full_name if inv.mechanic_task.mechanic else "N/A",
                            inv.mechanic_task.unique_task_id, inv.total_cost, inv.get_status_display(),
                            inv.created_at.strftime('%Y-%m-%d')])
    return _finalize_excel(wb, "invoice_reports")

def _export_mechanics_excel(mechanics):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mechanic Reports"
    _add_logo_and_title(ws, "Mechanic Report")
    _add_table(ws, ['Mechanic ID', 'Full Name', 'Total Tasks', 'Accepted Tasks', 'Completed Tasks', 'Invoices Sent',
                    'Pending Invoices', 'Approved Invoices', 'Paid Invoices'], mechanics,
               lambda mech: [mech['mechanic_id'], mech['full_name'], mech['total_tasks'], mech['accepted_tasks'],
                             mech['completed_tasks'], mech['invoices_sent'], mech['pending_invoices'],
                             mech['approved_invoices'], mech['paid_invoices']])
    return _finalize_excel(wb, "mechanic_reports")

def _export_drivers_excel(drivers):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Driver Reports"
    _add_logo_and_title(ws, "Driver Report")
    _add_table(ws, ['Driver ID', 'Full Name', 'Total Requests', 'Pending Requests', 'Approved Requests', 'Completed Requests'], drivers,
               lambda driver: [driver['driver_id'], f"{driver['user__first_name']} {driver['user__last_name']}".strip() or "N/A",
                               driver['total_requests'], driver['pending_requests'], driver['approved_requests'],
                               driver['completed_requests']])
    return _finalize_excel(wb, "driver_reports")

def _export_vehicles_excel(vehicles):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vehicle Reports"
    _add_logo_and_title(ws, "Vehicle Report")
    _add_table(ws, ['Number Plate', 'Make', 'Model', 'Total Repairs'], vehicles,
               lambda vehicle: [vehicle['number_plate'], vehicle['make'], vehicle['model'], vehicle['total_repairs']])
    return _finalize_excel(wb, "vehicle_reports")

def _add_logo_and_title(ws, title_text):
    logo_path = finders.find('managers/images/murangalogo.jpg')
    if logo_path:
        img = XLImage(logo_path)
        img.height, img.width = 60, 60
        ws.add_image(img, "A1")
    ws.merge_cells('C1:F1')
    ws['C1'] = f"Murang'a County - {title_text}"
    ws['C1'].font = Font(size=14, bold=True)
    ws['C1'].alignment = Alignment(horizontal="center")

def _add_table(ws, headers, data, row_func):
    ws.append([])
    ws.append(headers)
    header_fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    row_num = 4
    for item in data:
        row = row_func(item)
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        row_num += 1
    # Filter out MergedCell objects when setting column widths
    for col in ws.columns:
        for cell in col:
            if hasattr(cell, 'column_letter'):
                ws.column_dimensions[cell.column_letter].width = 20
                break  # Set width for the first valid cell in the column

def _finalize_excel(wb, filename_prefix):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response

def _export_maintenance_pdf(requests):
    return _export_pdf("Maintenance Report", requests, lambda req: [req.pk, req.driver.user.get_full_name() or req.driver.user.username,
                                                                   req.vehicle.number_plate, req.get_status_display(),
                                                                   req.submitted_at.strftime('%Y-%m-%d')],
                       ['Request ID', 'Driver', 'Vehicle', 'Status', 'Submitted Date'])

def _export_invoices_pdf(invoices):
    return _export_pdf("Invoice Report", invoices, lambda inv: [inv.id, inv.mechanic_task.mechanic.full_name if inv.mechanic_task.mechanic else "N/A",
                                                               inv.mechanic_task.unique_task_id, inv.total_cost,
                                                               inv.get_status_display(), inv.created_at.strftime('%Y-%m-%d')],
                       ['Invoice ID', 'Mechanic', 'Task ID', 'Total Cost', 'Status', 'Created Date'])

def _export_mechanics_pdf(mechanics):
    return _export_pdf("Mechanic Report", mechanics, lambda mech: [mech['mechanic_id'], mech['full_name'], mech['total_tasks'],
                                                                  mech['accepted_tasks'], mech['completed_tasks'], mech['invoices_sent'],
                                                                  mech['pending_invoices'], mech['approved_invoices'], mech['paid_invoices']],
                       ['Mechanic ID', 'Full Name', 'Total Tasks', 'Accepted Tasks', 'Completed Tasks', 'Invoices Sent',
                        'Pending Invoices', 'Approved Invoices', 'Paid Invoices'])

def _export_drivers_pdf(drivers):
    return _export_pdf("Driver Report", drivers, lambda driver: [driver['driver_id'],
                                                                f"{driver['user__first_name']} {driver['user__last_name']}".strip() or "N/A",
                                                                driver['total_requests'], driver['pending_requests'],
                                                                driver['approved_requests'], driver['completed_requests']],
                       ['Driver ID', 'Full Name', 'Total Requests', 'Pending Requests', 'Approved Requests', 'Completed Requests'])

def _export_vehicles_pdf(vehicles):
    return _export_pdf("Vehicle Report", vehicles, lambda vehicle: [vehicle['number_plate'], vehicle['make'], vehicle['model'],
                                                                   vehicle['total_repairs']],
                       ['Number Plate', 'Make', 'Model', 'Total Repairs'])

def _export_pdf(title_text, data, row_func, headers):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    logo_path = finders.find('managers/images/murangalogo.jpg')
    if logo_path:
        try:
            # Debug: Check if ReportlabImage is callable
            if callable(ReportlabImage):
                elements.append(ReportlabImage(logo_path, width=60, height=60))
            else:
                print(f"Debug: ReportlabImage is not callable, type: {type(ReportlabImage)}")
        except Exception as e:
            print(f"Error adding logo: {e}, skipping logo.")
    styles = getSampleStyleSheet()
    elements.append(Paragraph(f"<b>Murang'a County - {title_text}</b>", styles['Title']))
    elements.append(Spacer(1, 12))
    table_data = [headers]
    for item in data:
        table_data.append(row_func(item))
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFCC00')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey])
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{title_text.lower().replace(" ", "_")}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    return response

@login_required
def report_details(request, report_type, report_id):
    """Detailed report view for mechanics, drivers, vehicles, invoices, and maintenance."""
    details, requests_data, summary = None, [], {}
    if report_type == "mechanic":
        details = get_object_or_404(MechanicProfile, mechanic_id=report_id)
        requests_data = MaintenanceRequest.objects.filter(mechanic=details)

        total_tasks = MechanicTask.objects.filter(mechanic=details).count()
        completed_tasks = MechanicTask.objects.filter(mechanic=details, status="completed").count()
        invoices_sent = RepairInvoice.objects.filter(mechanic_task__mechanic=details).count()
        pending_invoices = RepairInvoice.objects.filter(mechanic_task__mechanic=details, status="pending").count()
        approved_invoices = RepairInvoice.objects.filter(mechanic_task__mechanic=details, status="approved").count()
        paid_invoices = RepairInvoice.objects.filter(mechanic_task__mechanic=details, status="paid").count()

        summary = {
            "total_tasks": total_tasks,
            "completed_tasks": completed_tasks,
            "invoices_sent": invoices_sent,
            "pending_invoices": pending_invoices,
            "approved_invoices": approved_invoices,
            "paid_invoices": paid_invoices,
        }
        details.type = "mechanic"

    elif report_type == "driver":
        details = get_object_or_404(DriverProfile, driver_id=report_id)
        requests_data = MaintenanceRequest.objects.filter(driver=details)

        summary = {
            "total_requests": requests_data.count(),
            "pending_requests": requests_data.filter(status="pending").count(),
            "approved_requests": requests_data.filter(status="approved").count(),
            "completed_requests": requests_data.filter(status="completed").count(),
        }
        details.type = "driver"

    elif report_type == "vehicle":
        details = get_object_or_404(Vehicle, number_plate=report_id)
        requests_data = MaintenanceRequest.objects.filter(vehicle=details)
        summary = {"total_repairs": requests_data.count()}
        details.type = "vehicle"

    elif report_type == "maintenance":
        details = get_object_or_404(MaintenanceRequest, pk=int(report_id))
        requests_data = [details]
        details.type = "maintenance"

    elif report_type == "invoice":
        details = get_object_or_404(RepairInvoice, pk=int(report_id))
        requests_data = MaintenanceRequest.objects.filter(assigned_mechanic=details.mechanic_task.mechanic)
        details.type = "invoice"

    else:
        return HttpResponseBadRequest("Invalid report type")

    return render(request, "managers/report_details.html", {
        "title": f"{report_type.capitalize()} Report",
        "details": details,
        "requests": requests_data,
        "summary": summary,
    })


@login_required
def settings(request):
    user = get_user(request)
    try:
        manager_profile = ManagerProfile.objects.get(user=user)
    except ManagerProfile.DoesNotExist:
        manager_profile = ManagerProfile.objects.create(user=user)

    if request.method == 'POST':
        if 'profile_submit' in request.POST:
            profile_form = ManagerProfileForm(request.POST, request.FILES, instance=user)
            if profile_form.is_valid():
                profile_form.save()
                if 'photo' in request.FILES:
                    manager_profile.photo = request.FILES['photo']
                    manager_profile.save()
                messages.success(request, "Profile updated successfully.")
                logger.info(f"User {user.username} updated their profile")
                return redirect('managers:settings')
            else:
                messages.error(request, "Please correct the errors in the profile form.")
        elif 'password_submit' in request.POST:
            password_form = PasswordChangeForm(user=user, data=request.POST)
            if password_form.is_valid():
                password_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, "Password updated successfully.")
                logger.info(f"User {user.username} updated their password")
                return redirect('managers:settings')
            else:
                messages.error(request, "Please correct the errors in the password form.")
    else:
        profile_form = ManagerProfileForm(instance=user)
        password_form = PasswordChangeForm(user=user)

    return render(request, 'managers/settings.html', {
        'greeting': get_greeting(),
        'user': user,
        'manager_profile': manager_profile,
        'profile_form': profile_form,
        'password_form': password_form,
    })

@login_required
def support(request):
    user = get_user(request)
    driver_requests = SupportRequest.objects.all().select_related('user').order_by('-created_at')
    mechanic_requests = MechanicSupportRequest.objects.all().select_related('mechanic__user').order_by('-created_at')

    status_filter = request.GET.get('status', 'all')
    type_filter = request.GET.get('type', 'all')

    if status_filter != 'all':
        driver_requests = driver_requests.filter(status=status_filter)
        mechanic_requests = mechanic_requests.filter(status=status_filter)
    
    if type_filter == 'driver':
        mechanic_requests = mechanic_requests.none()
    elif type_filter == 'mechanic':
        driver_requests = driver_requests.none()

    return render(request, 'managers/support_requests.html', {
        'greeting': get_greeting(),
        'user': user,
        'driver_requests': driver_requests,
        'mechanic_requests': mechanic_requests,
        'status_filter': status_filter,
        'type_filter': type_filter,
    })

@login_required
def add_response(request, request_type, request_id):
    user = get_user(request)
    if not (user.role == 'manager' or user.is_staff or user.is_superuser):
        logger.warning(f"Unauthorized attempt to respond to support request by {user.username} (role: {user.role})")
        messages.error(request, "You are not authorized to respond to support requests.")
        return redirect('core:profile_management')

    if request_type == 'driver':
        support_request = get_object_or_404(SupportRequest, pk=request_id)
    else:
        support_request = get_object_or_404(MechanicSupportRequest, pk=request_id)

    if request.method == 'POST':
        message = request.POST.get('message')
        if not message:
            messages.error(request, "Response message is required.")
            return redirect('managers:support_requests')
        
        SupportResponse.objects.create(
            driver_request=support_request if request_type == 'driver' else None,
            mechanic_request=support_request if request_type == 'mechanic' else None,
            responder=user,
            message=message
        )
        messages.success(request, "Response added successfully.")
        return redirect('managers:support_requests')

    return render(request, 'managers/support_response.html', {
        'greeting': get_greeting(),
        'request': support_request,
        'request_type': request_type,
    })

@login_required
def update_status(request, request_type, request_id):
    user = get_user(request)
    if not (user.role == 'manager' or user.is_staff or user.is_superuser):
        logger.warning(f"Unauthorized attempt to update support request status by {user.username} (role: {user.role})")
        messages.error(request, "You are not authorized to update support request statuses.")
        return redirect('core:profile_management')

    if request_type == 'driver':
        support_request = get_object_or_404(SupportRequest, pk=request_id)
    else:
        support_request = get_object_or_404(MechanicSupportRequest, pk=request_id)

    if request.method == 'POST':
        status = request.POST.get('status')
        if status in ['open', 'resolved', 'closed']:
            support_request.status = status
            support_request.save()
            messages.success(request, f"Support request marked as {status.capitalize()}.")
        else:
            messages.error(request, "Invalid status selected.")
        return redirect('managers:support_requests')

    return render(request, 'managers/support_requests.html', {
        'greeting': get_greeting(),
        'request': support_request,
        'request_type': request_type,
    })

@login_required
def drivers(request):
    return render(request, 'managers/drivers.html')

@login_required
def assignments(request):
    return render(request, 'managers/assignments.html')